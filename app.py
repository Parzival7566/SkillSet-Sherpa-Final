from flask import Flask, render_template, request, jsonify
import webbrowser
import json
from llamaapi import LlamaAPI
from img2table.ocr import EasyOCR
from img2table.document import PDF
import pandas as pd
import openpyxl
import os
import tempfile

app = Flask(__name__)
app.config["SECRET_KEY"] = "SECRET_KEY"

# Initialize LlamaAPI
llama = LlamaAPI('Insert_LLAMA_API_KEY_HERE')

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/upload-marksheet', methods=['POST'])
def upload_marksheet():
    if 'marksheet' in request.files:
        file = request.files['marksheet']
        # Ensure the uploads directory exists
        uploads_dir = os.path.join(os.getcwd(), 'uploads')
        os.makedirs(uploads_dir, exist_ok=True)
        filepath = os.path.join(uploads_dir, file.filename)
        file.save(filepath)
        
        # Call the process_marksheet function with the filepath
        process_marksheet(filepath)
        
        return jsonify({'filepath': filepath, 'message': 'File processed successfully'})
    return jsonify({'error': 'No file uploaded'})



@app.route('/bot_response', methods=['POST'])
def bot_response():
    global conversation_history
    try:
        # Check if there's a direct prompt in the request, if not, use user input
        prompt = request.json.get('prompt')
        if prompt:
            user_input = prompt
        else:
            user_input = request.json.get('user_input', '')
            conversation_history.append({"role": "user", "content": user_input})
        
        api_request_json = {
            "model": "llama-13b-chat",
            "messages": conversation_history if not prompt else [{"role": "system", "content": "start conversation"}, {"role": "user", "content": user_input}]
        }

        llama_response = llama.run(api_request_json)

        if llama_response:
            if not prompt:
                conversation_history.append({"role": "assistant", "content": json.dumps(llama_response.json(), indent=2)})
            return (json.dumps(llama_response.json(), indent=2))
        else:
            return jsonify({"error": "Failed to get API response."})
    except Exception as e:
        return jsonify({"error": str(e)})
    

# Function to save the table data to a CSV file
def save_table_to_csv(filename):
    try:
        # Read the xlsx file using openpyxl
        wb = openpyxl.load_workbook(filename, data_only=True)
        sheet = wb.active

        # Extract headers and data
        headers = [f'{cell.value}\n{sheet.cell(row=2, column=cell.column).value}' if cell.value else f'Unnamed_{i}' for i, cell in enumerate(sheet[1])]
        data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=3)]

        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)

        # Save DataFrame to CSV file
        df.to_csv('table_data.csv', index=False)
        print("CSV file saved successfully.")

        # Specify specific column names to include in the selection
        specific_column_names = ["SUBJECT", "TOTAL MARKS"]

        # Select columns based on specific names
        df_selected = df[[col for col in df.columns if any(spec_col.lower() in col.lower() for spec_col in specific_column_names)]]

        # Rename specific columns based on partial match
        column_rename_mapping = {}
        for col in df_selected.columns:
            for spec_col in specific_column_names:
                if spec_col.lower() in col.lower():
                    column_rename_mapping[col] = f"{spec_col}"

        df_selected = df_selected.rename(columns=column_rename_mapping)

        # Add the values of the column two columns left of the "GRADE" column to a new column named "TOTAL MARKS"
        grade_column_index = next((idx for idx, col_name in enumerate(df.columns) if "GRADE" in col_name), None)
        if grade_column_index is not None:
            two_columns_left_index = max(0, grade_column_index - 2)
            df_selected["TOTAL MARKS"] = df.iloc[:, two_columns_left_index]

        # Drop rows with NaN values in the 'SUBJECT' column
        df_selected = df_selected[df_selected["SUBJECT"].notna()]

        # Save DataFrame with selected and renamed columns to a new CSV file
        df_selected.to_csv('selected_columns_data.csv', index=False)
        print(df_selected)
        print("CSV file with selected and renamed columns saved successfully.")
        # Call the function to remove CSV files after processing the marksheet
        remove_csv_files()
    except Exception as e:
        print(f"Error: {e}")


# Function to remove CSV files after processing the marksheet
def remove_csv_files():
    try:
        os.remove('table_data.csv')
        os.remove('tables.xlsx')
        print("Cleaned up successfully.")
    except Exception as e:
        print(f"Error: {e}")

easyocr = EasyOCR(lang=["en"], kw={"gpu": False})
def process_marksheet(filepath):
    # Your OCR and data processing code here, adjusted to use 'filepath'
    # For example:
    pdf = PDF(src=filepath)
    pdf.to_xlsx('tables.xlsx',
                ocr=easyocr,
                implicit_rows=True,
                borderless_tables=False,
                min_confidence=50)
    save_table_to_csv('tables.xlsx')

questions = [
    ("Do you enjoy working with your hands and tools?", "Realistic (R)"),
    ("Are you naturally curious and enjoy exploring new ideas?", "Investigative (I)"),
    ("Do you have a passion for creative activities like drawing, writing, or music?", "Artistic (A)"),
    ("Do you find fulfillment in helping and working with others?", "Social (S)"),
    ("Are you ambitious and enjoy taking on leadership roles?", "Enterprising (E)"),
    ("Do you prefer working in structured and organized environments?", "Conventional (C)"),
    ("Are you interested in working outdoors or in natural settings?", "General (Mixed RIASEC)"),
    ("Do you prefer a fast-paced and dynamic work environment?", "General (Mixed RIASEC)"),
    ("Do you enjoy working with machinery and technology?", "Realistic (R)"),
    ("Do you like conducting experiments and solving complex problems?", "Investigative (I)"),
    ("Do you have a talent for artistic and creative expression?", "Artistic (A)"),
    ("Do you enjoy teaching or mentoring others?", "Social (S)"),
    ("Do you have strong negotiation and persuasion skills?", "Enterprising (E)"),
    ("Do you prefer a neat and structured workspace?", "Conventional (C)"),
    ("Do you like spending time in nature and outdoor activities?", "General (Mixed RIASEC)"),
    ("Do you thrive in high-pressure and competitive situations?", "General (Mixed RIASEC)"),
    ("Are you mechanically inclined and good at fixing things?", "Realistic (R)"),
    ("Do you enjoy analyzing data and conducting research?", "Investigative (I)"),
    ("Are you skilled in playing musical instruments or creating art?", "Artistic (A)"),
    ("Do you excel in teamwork and collaboration?", "Social (S)")
]

user_responses = {}

@app.route('/aptitude')
def aptitude():
    return render_template('aptitude.html')

@app.route('/index2')
def index2():
    return render_template('index_2.html')

# add a new route to return the response.txt filw to the server
@app.route('/response')
def response():
    return render_template('response.txt')

@app.route('/submit_responses', methods=['POST'])
def submit_responses():
    try:
        responses = request.json  # This is already a dictionary
        num_questions = 20  # Update this number if needed
        
        # Instead of requiring all questions to be answered, handle cases where some might be missing
        if len(responses) > num_questions:
            return jsonify({"error": "Received more responses than expected."}), 400
        
        # Calculate RIASEC scores
        riasec_scores = calculate_riasec_scores(responses)
        
        # Prepare the response with a success message and redirection link
        response = {
            "message": "Your responses have been submitted successfully!",
            "redirect_url": "/index2"  # Replace with your desired redirection path
        }
        
        # Return the JSON response
        return jsonify(response)
    except Exception as e:
        return jsonify({"error": str(e)})


def calculate_riasec_scores(responses):
    if len(responses) != len(questions):
        return {"error": "Number of responses does not match the number of questions."}

    # Convert responses to scores
    scores = [int(score) for score in responses.values()]
    total_score = sum(scores)
    if total_score == 0:  # Prevent division by zero
        return {"error": "Total score cannot be zero."}

    normalized_scores = {}
    for (_, category), score in zip(questions, scores):
        category_key = category.split()[0]
        normalized_scores[category_key] = normalized_scores.get(category_key, 0) + score

    # Normalize scores
    for key in normalized_scores:
        normalized_scores[key] = (normalized_scores[key] / total_score) * 100
    
    print(normalized_scores)

    # Check if marksheet data is available
    if os.path.exists('selected_columns_data.csv'):
        # Read the marksheet data from the CSV file
        marksheet_data = pd.read_csv('selected_columns_data.csv')
        
        # Create the temporary prompt file
        temp_prompt_file = create_temp_prompt(marksheet_data.to_dict(), normalized_scores)
        
        # Send the prompt to the llama API
        with open(temp_prompt_file, 'r') as file:
            prompt = file.read()
        
        api_request_json = {
            "model": "llama-13b-chat",
            "messages": [{"role": "system", "content": "start conversation"}, {"role": "user", "content": prompt}]
        }

        llama_response = llama.run(api_request_json)

        if llama_response:
            # Save the API response to a text file in the templates directory
            response_text = json.dumps(llama_response.json(), indent=2)
            with open('templates/response.txt', 'w') as file:
                file.write(response_text)
            
            # Remove the temporary prompt file
            os.remove(temp_prompt_file)
        else:
            print("Failed to get API response.")
    
    return normalized_scores

# ... (remaining code remains the same)

def create_temp_prompt(marks_data, aptitude_data):
    with tempfile.NamedTemporaryFile(mode='w+', delete=False) as tmp:
        prompt = "Hi, I am a student of 10th class and I want to know my future career options. Please answer the following questions based on my marks and aptitude test.\n"
        prompt += f"Marks Data: {marks_data}\n"
        prompt += f"Aptitude Data: {aptitude_data}\n"
        prompt += "Based on my marks and aptitude test, what are the best career options for me?\n For achieveing those career options, what are the streams that I should take in the coming 2 years?\n Justify your answer.\n"
        tmp.write(prompt)
        print(prompt)
        return tmp.name  # Returns the path to the temporary file

conversation_history = []

if __name__ == '__main__':
    webbrowser.open('http://127.0.0.1:5000')
    app.run(debug = True)
