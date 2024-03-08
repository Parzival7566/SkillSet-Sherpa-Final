from img2table.ocr import EasyOCR
from img2table.document import PDF
import pandas as pd
import openpyxl

easyocr = EasyOCR(lang=["en"], kw={"gpu": False})
filepath = "uploads/Class X Original Marksheet.pdf"
def process_marksheet(filepath):
    # Your OCR and data processing code here, adjusted to use 'filepath'
    # For example:
    pdf = PDF(src=filepath)
    pdf.to_xlsx('tables.xlsx',
                ocr=easyocr,
                implicit_rows=True,
                borderless_tables=False,
                min_confidence=50)
    print("Done")
    save_table_to_csv('tables.xlsx')

# Function to save the table data to a CSV file
def save_table_to_csv(filename):
    try:
        # Read the xlsx file using openpyxl
        wb = openpyxl.load_workbook(filename, data_only=True)
        sheet = wb.active

        # Extract headers and data
        headers = [f'{cell.value}\n{sheet.cell(row=2, column=cell.column).value}' if cell.value else f'Unnamed_{i}' for i, cell in enumerate(sheet[1])]
        data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=3)]

        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)

        # Save DataFrame to CSV file
        df.to_csv('table_data.csv', index=False)
        print("CSV file saved successfully.")

        # Specify specific column names to include in the selection
        specific_column_names = ["SUBJECT", "TOTAL MARKS"]

        # Select columns based on specific names
        df_selected = df[[col for col in df.columns if any(spec_col.lower() in col.lower() for spec_col in specific_column_names)]]

        # Rename specific columns based on partial match
        column_rename_mapping = {}
        for col in df_selected.columns:
            for spec_col in specific_column_names:
                if spec_col.lower() in col.lower():
                    column_rename_mapping[col] = f"{spec_col}"

        df_selected = df_selected.rename(columns=column_rename_mapping)
        print(df)

        # Add the values of the column two columns left of the "GRADE" column to a new column named "TOTAL MARKS"
        grade_column_index = next((idx for idx, col_name in enumerate(df.columns) if "GRADE" in col_name), None)
        if grade_column_index is not None:
            two_columns_left_index = max(0, grade_column_index - 2)
            df_selected["TOTAL MARKS"] = df.iloc[:, two_columns_left_index]

        # Drop rows with NaN values in the 'SUBJECT' column
        df_selected = df_selected[df_selected["SUBJECT"].notna()]

        # Save DataFrame with selected and renamed columns to a new CSV file
        df_selected.to_csv('selected_columns_data.csv', index=False)
        print("CSV file with selected and renamed columns saved successfully.")
    except Exception as e:
        print(f"Error: {e}")

process_marksheet(filepath)